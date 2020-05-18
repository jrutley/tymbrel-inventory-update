import os
import argparse
from dotenv import load_dotenv
from typing import List
from openpyxl import load_workbook

from dto import Product
from web import setup, insert_products
load_dotenv()

def assignValues(headerTuple) -> dict:
  valueDict = dict()
  for value in headerTuple:
    valueDict[value] = headerTuple.index(value)
  return valueDict

if __name__ == "__main__":

  parser = argparse.ArgumentParser(description='Update inventory in Tymbrel')
  parser.add_argument('filename', type=str, nargs=1, default='Inventory.xlsx',
                      help='the Excel file containing inventory items')

  args = parser.parse_args()
  importCsv: str = args.filename[0]
  
  if not os.path.exists(importCsv):
    print(f"Couldn't find {importCsv} in the current folder")
    exit(1)

  workbook = load_workbook(filename=importCsv)
  sheet = workbook[workbook.sheetnames[0]]
  #sheet = workbook.active

  for value in sheet.iter_rows(min_row=1,max_row=1,values_only=True):
    # value is a Tuple
    # assign the positions of the tuple
    columns = assignValues(value)
  products = []
  for value in sheet.iter_rows(min_row=2,values_only=True):
    quantity = value[columns["QUANTITY"]]
    if quantity == '' or quantity is None:
      quantity = '0'
    products.append(Product(sku = value[columns["SKU"]],
                            product = value[columns["PRODUCT"]],
                            skuName = value[columns["SKU NAME"]],
                            productBasePrice = value[columns["PRODUCT BASE PRICE"]],
                            skuPricePerUnit = value[columns["SKU PRICE PER UNIT"]],
                            categories = value[columns["CATEGORIES"]],
                            quantity = int(quantity)))


  session = setup()
  products = insert_products(products, session)
  error_count = 0
  with open("errors.txt", "w") as error_file:
    for p in products:
      if isinstance(p, Exception):
        error_count += 1
        arg_str = str(type(p))
        for arg in p.args:
          arg_str += f" \"{arg}\""
        arg_str += os.linesep
        error_file.write(arg_str)
        continue
    print("Done")
  if error_count > 0:
    print(f"Exited with {error_count} errors. Open errors.txt to get the details")